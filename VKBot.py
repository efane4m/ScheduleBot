import requests
import openpyxl
import os.path
import pickle
from bs4 import BeautifulSoup

import datetime
from datetime import date
from datetime import timedelta
from datetime import datetime

import vk_api
from vk_api.longpoll import VkLongPoll, VkEventType
from vk_api.keyboard import VkKeyboard, VkKeyboardColor
from vk_api.utils import get_random_id

import pyowm
from pyowm.utils.config import get_default_config
from pyowm.utils import formatting
from pyowm.weatherapi25.weather import Weather

def get_schedule(new_data = False):
    if os.path.exists('schedule.pickle') and not new_data:
        with open('schedule.pickle', 'rb') as f:
            groups, teachers = pickle.load(f)
            return groups, teachers
    page = requests.get("https://www.mirea.ru/schedule/")
    soup = BeautifulSoup(page.text, "html.parser")
    result = soup.find("div", {"class": "schedule"}).\
             find(string = "Институт информационных технологий").\
             find_parent("div").\
             find_parent("div").\
             findAll("a", {"class": "uk-link-toggle"})
    groups = {}
    teachers = {}
    for x in result:
        link = str(x)[str(x).find('https://'):str(x).find('" target=')]
        book_name = link[link.rfind('/') + 1:]
        if book_name[0:3] != "IIT":
            continue
        file = open(book_name, "wb")
        file.write(requests.get(link).content)
        file.close()
        book = openpyxl.load_workbook(book_name)
        sheet = book.active
        for y in range(1, sheet.max_column + 1):
            group = sheet.cell(row = 2, column = y).value
            if group and str(group).count('-') == 2:
                week_1 = []
                week_2 = []
                week = [week_1, week_2]
                for z in range (2):
                    for i in range(7):
                        day = []
                        for j in range(7):
                            if sheet.cell(row = 4 + z + i * 14 + j * 2, column = y).value:
                                obj = [sheet.cell(row = 4 + z + i * 14 + j * 2, column = y).value,
                                       sheet.cell(row = 4 + z + i * 14 + j * 2, column = y + 1).value,
                                       sheet.cell(row = 4 + z + i * 14 + j * 2, column = y + 2).value,
                                       sheet.cell(row = 4 + z + i * 14 + j * 2, column = y + 3).value]
                                for k in range(4):
                                    if obj[k].count("\n") != 0:
                                        obj[k] = obj[k][:obj[k].find("\n")]
                                if obj[2].count(",") != 0:
                                    obj[2] = obj[2][:obj[2].find(",")]
                                if not obj[2] in teachers and obj[2] != "":
                                    teacher = [[[str(j+1) + ") -" for j in range(7)] for i in range(7)] for z in range(2)]
                                    teachers[obj[2]] = teacher
                                if obj[2] != "":
                                    obj_ = obj.copy()
                                    obj_[2] = group
                                    teachers[obj[2]][z][i][j] = str(j+1) + ") " + ", ".join(map(str, obj_))
                            else:
                                obj = ['-']
                            day.append(str(j+1) + ") " + ", ".join(map(str, obj)))
                        week[z].append(day)
                groups[group] = week
    with open('schedule.pickle', 'wb') as f:
        pickle.dump([groups, teachers], f)
    print("Ready")
    return groups, teachers

def get_schedule_by_day(groups, group, dt):
    first_day = date(year = 2024, month = 9, day = 2)
    days = (dt - first_day).days % 14
    if days < 7:
        return groups[group][0][days]
    else:
        return groups[group][1][days - 7]

def get_schedule_by_week_day(groups, group, day):
    week_days = {"понедельник": 1,
                 "вторник": 2,
                 "среда": 3,
                 "четверг": 4,
                 "пятница": 5,
                 "суббота": 6,
                 "воскресенье": 7}
    return [groups[group][0][week_days[day] - 1],
            groups[group][1][week_days[day] - 1]]

def get_schedule_by_week(groups, group, dt):
    first_day = date(year = 2024, month = 9, day = 2)
    days = (dt - first_day).days % 14
    if days < 7:
        return groups[group][0]
    else:
        return groups[group][1]

def get_weather(weather_manager, day: datetime):
    weather_1 = weather_manager.one_call_history(55.753, 37.621, dt = formatting.to_UNIXtime(day) - 100).forecast_hourly if day <= datetime.now() + timedelta(minutes = 1) else []
    weather_2 = weather_manager.one_call(55.753, 37.621).forecast_hourly
    weather_all = list(set(weather_1 + weather_2))
    weather_all.sort(key = lambda x: x.ref_time)
    return [weather for weather in weather_all if datetime.utcfromtimestamp(weather.ref_time).day == day.day]

def weather_to_text(weather: Weather):
    weather_icons = {"ясно": '☀️',
                 'небольшая облачность': '🌤',
                 'облачно с прояснениями': '🌤',
                 'переменная облачность': '☁',
                 'пасмурно': '☁',
                 'небольшой дождь': '🌦',
                 'дождь': '🌧',
                 'сильный дождь': '🌧'}
    text = f"{weather_icons[weather.detailed_status]}\n{weather.detailed_status}, температура {weather.temperature('celsius')['temp']}°С\n" \
           f"Давление: {int(weather.pressure['press'] / 1.33322)} мм ртутного столба, влажность: {weather.humidity}%\n" \
           f"Ветер {weather.wind()['speed']} м\\с\n"
    return text

def main():
    vk_session = vk_api.VkApi(token = 'YOU_VK_TOKEN')
    vk = vk_session.get_api()
    presets = get_default_config()
    presets['language'] = 'ru'
    owm = pyowm.OWM('YOU_API_TOKEN', presets)
    weather_manager = owm.weather_manager()
    keyboard_main = VkKeyboard(one_time = False)
    keyboard_main.add_button('на сегодня', color=VkKeyboardColor.POSITIVE)
    keyboard_main.add_button('на завтра', color=VkKeyboardColor.NEGATIVE)
    keyboard_main.add_line()
    keyboard_main.add_button('на эту неделю', color=VkKeyboardColor.SECONDARY)
    keyboard_main.add_button('на следующую неделю', color=VkKeyboardColor.SECONDARY)
    keyboard_main.add_line()
    keyboard_main.add_button('какая неделя?')
    keyboard_main.add_button('какая группа?')
    keyboard_short = VkKeyboard(one_time = False)
    keyboard_short.add_button('на сегодня', color=VkKeyboardColor.POSITIVE)
    keyboard_short.add_button('на завтра', color=VkKeyboardColor.NEGATIVE)
    keyboard_short.add_line()
    keyboard_short.add_button('на эту неделю', color=VkKeyboardColor.SECONDARY)
    keyboard_short.add_button('на следующую неделю', color=VkKeyboardColor.SECONDARY)
    keyboard_weather = VkKeyboard(one_time = False)
    keyboard_weather.add_button('сейчас', color=VkKeyboardColor.SECONDARY)
    keyboard_weather.add_button('сегодня', color=VkKeyboardColor.POSITIVE)
    keyboard_weather.add_button('завтра', color=VkKeyboardColor.POSITIVE)
    keyboard_weather.add_line()
    keyboard_weather.add_button('на 5 дней', color=VkKeyboardColor.POSITIVE)
    week_days = {"понедельник": 1,
                 "вторник": 2,
                 "среда": 3,
                 "четверг": 4,
                 "пятница": 5,
                 "суббота": 6,
                 "воскресенье": 7}
    weather_icons = {"ясно": '☀️',
                 'небольшая облачность': '🌤',
                 'облачно с прояснениями': '🌤',
                 'переменная облачность': '☁',
                 'пасмурно': '☁',
                 'небольшой дождь': '🌦',
                 'дождь': '🌧',
                 'сильный дождь': '🌧'}
    groups, teachers = get_schedule()
    users_current = {}
    users_groups = {}
    names = []
    longpoll = VkLongPoll(vk_session)
    weather = [get_weather(weather_manager, datetime.today())] + [get_weather(weather_manager, datetime.today() + timedelta(days = 1))]
    for event in longpoll.listen():
        if event.type == VkEventType.MESSAGE_NEW and event.text and event.to_me:
            txt = event.text.lower()
            words = txt.split()
            keyboard = keyboard_main.get_keyboard()
            if event.user_id not in users_groups:
                users_groups[event.user_id] = "ИКБО-25-22"
                users_current[event.user_id] = "ИКБО-25-22"
            if txt == 'начать' or txt == 'start':
                message = '''Доступные команды:
1) [номер группы] - установка номера группы
2) "на сегодня" - получение расписания на сегодня
3) "на завтра" - получение расписание на завтра
4) "какая неделя?" - получение текущей недели
5) "на эту неделю" - получение расписания на эту неделю
6) "на следующую неделю" - получение расписания на следующую неделю
7) "какая группа?" - получение текущей группы
8) "Найти [фамилия]" - получение расписания преподователя
9) "погода" - получение информации о погоде в Москве'''
                keyboard = 0
            elif txt.upper() in groups:
                users_groups[event.user_id] = txt.upper()
                message = "Я запомнил, что ты из группы " + txt.upper()
                keyboard = 0
            elif len(words) == 2 and words[1] in week_days and users_current[event.user_id] in groups:
                schedule = get_schedule_by_week_day(groups, users_current[event.user_id], words[1])
                message = "\n".join(["Расписание на " + words[1] + " группы " + users_current[event.user_id],
                                     "Нечетная неделя:", *schedule[0],
                                     "Четная неделя:", *schedule[1]])
                users_current[event.user_id] = users_groups[event.user_id]
            elif len(words) == 2 and words[1].upper() in groups:
                message = "Показать расписание группы " + words[1] + "..."
                users_current[event.user_id] = words[1].upper()
            elif len(words) == 3 and words[1] in week_days and words[2].upper() in groups:
                users_current[event.user_id] = words[2].upper()
                schedule = get_schedule_by_week_day(groups, users_current[event.user_id], words[1])
                message = "\n".join(["Расписание на " + words[1] + " группы " + users_current[event.user_id],
                                     "Нечетная неделя:", *schedule[0],
                                     "Четная неделя:", *schedule[1]])
                users_current[event.user_id] = users_groups[event.user_id]
            elif txt == "на сегодня" and users_current[event.user_id] in groups:
                schedule = get_schedule_by_day(groups, users_current[event.user_id], date.today())
                message = "\n".join(["Расписание на сегодня группы " + users_current[event.user_id], *schedule])
                users_current[event.user_id] = users_groups[event.user_id]
            elif txt == "на завтра" and users_current[event.user_id] in groups:
                schedule = get_schedule_by_day(groups, users_current[event.user_id], date.today() + timedelta(days = 1))
                message = "\n".join(["Расписание на завтра группы " + users_current[event.user_id], *schedule])
                users_current[event.user_id] = users_groups[event.user_id]
            elif txt == "на эту неделю" and users_current[event.user_id] in groups:
                schedule = get_schedule_by_week(groups, users_current[event.user_id], date.today())
                message = ["Расписание на эту неделю группы " + users_current[event.user_id]]
                for key in week_days:
                    message.append(key)
                    message += schedule[week_days[key] - 1]
                message = "\n".join(message)
                users_current[event.user_id] = users_groups[event.user_id]
            elif txt == "на следующую неделю" and users_current[event.user_id] in groups:
                schedule = get_schedule_by_week(groups, users_current[event.user_id], date.today() + timedelta(days=7))
                message = ["Расписание на следующую неделю группы " + users_current[event.user_id]]
                for key in week_days:
                    message.append(key)
                    message += schedule[week_days[key] - 1]
                message = "\n".join(message)
                users_current[event.user_id] = users_groups[event.user_id]
            elif txt == 'какая неделя?':
                message = "Идет " + str((date.today() - date(year = 2024, month = 9, day = 2)).days // 7 + 1) + " неделя"
            elif txt == 'какая группа?':
                message = "Показываю расписание группы " + users_current[event.user_id]
            elif words[0] == "найти":
                names = []
                for teacher in teachers:
                    if words[1] in str(teacher).lower():
                        names += [str(teacher)]
                if len(names) > 1:
                    keyboard_teachers = VkKeyboard(one_time = True)
                    for x in range(5):
                        if x >= len(names):
                            break
                        if x > 0:
                            keyboard_teachers.add_line()
                        keyboard_teachers.add_button(str(names[x]), color=VkKeyboardColor.SECONDARY)
                    message = "Выберите преподавателя"
                    keyboard = keyboard_teachers.get_keyboard()
                else:
                    users_current[event.user_id] = names[0]
                    message = "Показать расписание преподавателя " + names[0]
                    keyboard = keyboard_short.get_keyboard()
            elif event.text in names:
                users_current[event.user_id] = event.text
                message = "Показать расписание преподавателя " + event.text
                keyboard = keyboard_short.get_keyboard()
            elif txt == "на сегодня" and users_current[event.user_id] in teachers:
                schedule = get_schedule_by_day(teachers, users_current[event.user_id], date.today())
                message = "\n".join(["Расписание на сегодня группы " + users_current[event.user_id], *schedule])
                users_current[event.user_id] = users_groups[event.user_id]
            elif txt == "на завтра" and users_current[event.user_id] in teachers:
                schedule = get_schedule_by_day(teachers, users_current[event.user_id], date.today() + timedelta(days = 1))
                message = "\n".join(["Расписание на завтра группы " + users_current[event.user_id], *schedule])
                users_current[event.user_id] = users_groups[event.user_id]
            elif txt == "на эту неделю" and users_current[event.user_id] in teachers:
                schedule = get_schedule_by_week(teachers, users_current[event.user_id], date.today())
                message = ["Расписание на эту неделю группы " + users_current[event.user_id]]
                for key in week_days:
                    message.append(key)
                    message += schedule[week_days[key] - 1]
                message = "\n".join(message)
                users_current[event.user_id] = users_groups[event.user_id]
            elif txt == "на следующую неделю" and users_current[event.user_id] in teachers:
                schedule = get_schedule_by_week(teachers, users_current[event.user_id], date.today() + timedelta(days=7))
                message = ["Расписание на следующую неделю группы " + users_current[event.user_id]]
                for key in week_days:
                    message.append(key)
                    message += schedule[week_days[key] - 1]
                message = "\n".join(message)
                users_current[event.user_id] = users_groups[event.user_id]
            elif txt == "погода":
                message = "Показать погоду в Москве"
                keyboard = keyboard_weather.get_keyboard()
            elif txt == "сейчас":
                weather_now = weather_manager.weather_at_place("Moscow").weather
                message = f"{weather_icons[weather_now.detailed_status]}\n{weather_now.detailed_status}, температура {weather_now.temperature('celsius')['temp']}°С\n" \
                       f"Давление: {int(weather_now.pressure['press'] / 1.33322)} мм ртутного столба, влажность: {weather_now.humidity}%\n" \
                       f"Ветер {weather_now.wind()['speed']} м\\с\n"
                keyboard = keyboard_weather.get_keyboard()
            elif txt == "сегодня":
                message = f'''Погода в Москве сегодня
                /{weather[0][9].temperature('celsius')['temp']}°С // {weather[0][14].temperature('celsius')['temp']}°С // {weather[0][17].temperature('celsius')['temp']}°С // {weather[0][23].temperature('celsius')['temp']}°С /
                УТРО  {weather_to_text(weather[0][9])}
                ДЕНЬ  {weather_to_text(weather[0][14])}
                ВЕЧЕР {weather_to_text(weather[0][17])}
                НОЧЬ  {weather_to_text(weather[0][23])}'''
                keyboard = keyboard_weather.get_keyboard()
            elif txt == "завтра":
                message = f'''Погода в Москве завтра
                /{weather[1][9].temperature('celsius')['temp']}°С // {weather[1][14].temperature('celsius')['temp']}°С // {weather[1][17].temperature('celsius')['temp']}°С // {weather[1][23].temperature('celsius')['temp']}°С /
                УТРО  {weather_to_text(weather[1][9])}
                ДЕНЬ  {weather_to_text(weather[1][14])}
                ВЕЧЕР {weather_to_text(weather[1][17])}
                НОЧЬ  {weather_to_text(weather[1][23])}'''
                keyboard = keyboard_weather.get_keyboard()
            elif txt == "на 5 дней":
                days = weather_manager.one_call(55.752546, 37.621193).forecast_daily[:5]
                message = f'''Погода в Москве с {str(datetime.today())[:10]} по {str(datetime.today() + timedelta(days=+4))[:10]}
                {weather_icons[days[0].detailed_status]} {weather_icons[days[1].detailed_status]} {weather_icons[days[2].detailed_status]} {weather_icons[days[3].detailed_status]} {weather_icons[days[4].detailed_status]}
                / {round((days[0].temp['day']) - 273.15, 2)}°С // {round((days[1].temp['day']) - 273.15, 2)}°С // {round((days[2].temp['day']) - 273.15, 2)}°С // {round((days[3].temp['day']) - 273.15, 2)}°С // {round((days[4].temp['day']) - 273.15, 2)}°С / ДЕНЬ
                / {round((days[0].temp['night']) - 273.15, 2)}°С // {round((days[1].temp['night']) - 273.15, 2)}°С // {round((days[2].temp['night']) - 273.15, 2)}°С // {round((days[3].temp['night']) - 273.15, 2)}°С // {round((days[4].temp['night']) - 273.15, 2)}°С / НОЧЬ'''
                keyboard = keyboard_weather.get_keyboard()
            else:
                message = "Неизвестная команда"
            vk.messages.send(
                user_id = event.user_id,
                random_id = get_random_id(),
                message = message,
                keyboard = keyboard)
if __name__ == '__main__':
    main()
